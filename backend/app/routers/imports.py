"""엑셀 업로드/DB 반영 라우터.

중요 원칙
- 전체면허자현황은 반드시 '개인' + '택배' 시트만 읽는다.
- 업체세분/선진물류/업체/차량집계는 전체자명단 업로드에서 절대 사용하지 않는다.
- 2026미수금은 반드시 '2026년회비내역' 시트만 읽는다.
- 기존 데이터는 일반 업로드에서 삭제하지 않는다. 초기화 버튼을 눌렀을 때만 misu_* 업무 테이블을 비운다.
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Any

import pandas as pd
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import delete, func, or_, select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from ..billing import charge_item, monthly_charge, next_month_ym
from ..database import get_db
from ..models import Closure, Deposit, Member, MemberHistory, Payment, Pending, ReceivableItem

router = APIRouter(prefix="/api/import", tags=["import"])

SIGUN = [
    "춘천시", "원주시", "강릉시", "동해시", "태백시", "속초시", "삼척시",
    "홍천군", "횡성군", "영월군", "평창군", "정선군", "철원군", "화천군",
    "양구군", "인제군", "고성군", "양양군",
]

STANDARD_MEMBER_COLUMNS = [
    "지역", "회원구분", "관리번호", "차량번호", "성명", "주민등록번호", "주소", "전화번호", "핸드폰",
    "인가일자", "가입일자", "자격증명 발급일자", "자격증명 발급번호", "운전면허증번호",
    "차종", "유종", "사업자등록번호", "소속업체", "공문 주소", "대리인", "구조변경",
    "비고", "전화 메모", "가입여부", "부과구분",
]

ARREARS_PREVIEW_COLUMNS = [
    "지역", "계정", "비고", "차량번호", "성명", "대수", "이월금",
    "기준월", "미수금액",
    "1월 미수금", "2월 미수금", "3월 미수금", "4월 미수금", "5월 미수금", "6월 미수금",
    "7월 미수금", "8월 미수금", "9월 미수금", "10월 미수금", "11월 미수금", "12월 미수금",
]


def _clean(v: Any) -> str:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.strip()


def _norm_col(v: Any) -> str:
    return re.sub(r"\s+", "", _clean(v))


def _person_name(v: Any) -> str:
    return re.sub(r"\s+", "", _clean(v))


def _clip(v: Any, n: int) -> str | None:
    s = _clean(v)
    return s[:n] if s else None


def _money(v: Any) -> int:
    s = _clean(v).replace("\u3000", "")
    if not s or s in {"-", "–", "—"}:
        return 0
    s = re.sub(r"[^0-9\-]", "", s)
    try:
        return max(0, int(s or 0))
    except Exception:
        return 0


def _parse_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _clean(v)
    if not s or s.lower() == "x":
        return None
    try:
        d = pd.to_datetime(s, errors="coerce")
        if pd.notna(d):
            return d.date()
    except Exception:
        pass
    digits = re.sub(r"\D", "", s)
    try:
        if len(digits) == 8:
            return date(int(digits[:4]), int(digits[4:6]), int(digits[6:8]))
        if len(digits) == 6:
            yy = int(digits[:2])
            year = 2000 + yy if yy < 80 else 1900 + yy
            return date(year, int(digits[2:4]), int(digits[4:6]))
    except Exception:
        return None
    return None


def _json(v: Any) -> Any:
    if isinstance(v, (date, datetime)):
        return v.isoformat()
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    return _clean(v)


def _sigun_from_text(text: str) -> str:
    t = _clean(text).replace(" ", "")
    for s in SIGUN:
        if s.replace(" ", "") in t or s[:-1] in t:
            return s
    return "미분류"


def _vehicle_last4(vehicle: str) -> str:
    nums = re.findall(r"\d+", _clean(vehicle))
    if not nums:
        return ""
    return nums[-1][-4:]


def _valid_vehicle(vehicle: str) -> bool:
    v = _clean(vehicle)
    if not v or "?" in v:
        return False
    return len(_vehicle_last4(v)) >= 4


def _birth_year(rrn: str) -> int | None:
    s = re.sub(r"\D", "", _clean(rrn))
    if len(s) < 7:
        return None
    yy = int(s[:2])
    marker = s[6]
    if marker in {"1", "2", "5", "6"}:
        return 1900 + yy
    if marker in {"3", "4", "7", "8"}:
        return 2000 + yy
    return 1900 + yy if yy > 30 else 2000 + yy


def _membership_from(row: dict[str, Any]) -> str:
    val = _clean(row.get("가입여부"))
    join = _clean(row.get("가입일자"))
    negative = {"x", "X", "×", "미가입", "비가입", "없음", "-"}
    if val in negative or "미가입" in val:
        return "협회미가입"
    if "가입" in val and "미가입" not in val:
        return "협회가입"
    if join and join not in negative and "미가입" not in join:
        return "협회가입"
    return "협회미가입"


def _next_member_id_from_no(no: int) -> str:
    return f"M{no:05d}"


def _current_max_member_no(db: Session) -> int:
    ids = db.scalars(select(Member.id).where(Member.id.like("M%"))).all()
    max_no = 0
    for mid in ids:
        m = re.search(r"(\d+)$", str(mid or ""))
        if m:
            max_no = max(max_no, int(m.group(1)))
    return max_no


def _make_mgmt_no(raw: str, fallback_no: int) -> str:
    s = _clean(raw)
    return (s or f"신26-{fallback_no:03d}")[:16]


def _find_existing_member(db: Session, name: str, vehicle_no: str) -> Member | None:
    name = _person_name(name)
    last4 = _vehicle_last4(vehicle_no)
    exact = db.scalar(select(Member).where(Member.vehicle_no == vehicle_no).limit(1))
    if exact:
        return exact
    if last4 and name:
        candidates = db.scalars(select(Member).where(Member.name == name)).all()
        for c in candidates:
            if _vehicle_last4(c.vehicle_no) == last4:
                return c
    return None


def _find_member_for_arrears(db: Session, name: str, vehicle_no: str) -> Member | None:
    return _find_existing_member(db, name, vehicle_no)


def _read_workbook(file_bytes: bytes):
    return load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True, keep_links=False)


def _header_map(ws, max_col: int = 80) -> tuple[list[str], dict[str, int]]:
    row = next(ws.iter_rows(min_row=1, max_row=1, max_col=max_col, values_only=True))
    headers: list[str] = []
    mapping: dict[str, int] = {}
    for idx, v in enumerate(row, start=1):
        h = _clean(v)
        if not h:
            continue
        h = re.sub(r"\s+", " ", h.replace("\n", " ")).strip()
        headers.append(h)
        mapping[_norm_col(h)] = idx
    return headers, mapping


def _get(row: tuple, mapping: dict[str, int], *keys: str) -> Any:
    for key in keys:
        idx = mapping.get(_norm_col(key))
        if idx is not None and idx - 1 < len(row):
            return row[idx - 1]
    return None


def _iter_license_rows(file_bytes: bytes, preview_limit: int | None = None) -> list[dict[str, Any]]:
    wb = _read_workbook(file_bytes)
    rows: list[dict[str, Any]] = []
    for sheet_name, member_type in (("개인", "개인"), ("택배", "택배")):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        _, m = _header_map(ws, max_col=80)
        for row in ws.iter_rows(min_row=2, max_col=80, values_only=True):
            vehicle = _clean(_get(row, m, "차량번호"))
            name = _person_name(_get(row, m, "성명", "성 명", "성    명"))
            if not _valid_vehicle(vehicle) or not name:
                continue
            region = _clean(_get(row, m, "지역")) or _sigun_from_text(_clean(_get(row, m, "주소")))
            membership = _membership_from({
                "가입여부": _get(row, m, "가입여부"),
                "가입일자": _get(row, m, "가입일자"),
            })
            row_out = {
                "지역": region or "미분류",
                "회원구분": member_type,
                "관리번호": _clean(_get(row, m, "관리번호")),
                "차량번호": vehicle,
                "성명": name,
                "주민등록번호": _clean(_get(row, m, "주민등록번호")),
                "주소": _clean(_get(row, m, "주소", "주 소")),
                "전화번호": _clean(_get(row, m, "전화번호")),
                "핸드폰": _clean(_get(row, m, "핸드폰", "핸 드 폰")),
                "인가일자": _json(_get(row, m, "인가일자")),
                "가입일자": _json(_get(row, m, "가입일자")),
                "자격증명 발급일자": _json(_get(row, m, "자격증명 발급일자", "자격증명 발급일", "자격증명\n발급일자")),
                "자격증명 발급번호": _clean(_get(row, m, "자격증명 발급번호", "자격증명\n발급번호")),
                "운전면허증번호": _clean(_get(row, m, "운전면허증번호")),
                "차종": _clean(_get(row, m, "차종")),
                "유종": _clean(_get(row, m, "유종")),
                "사업자등록번호": _clean(_get(row, m, "사업자등록번호")),
                "소속업체": _clean(_get(row, m, "소속업체")),
                "공문 주소": _clean(_get(row, m, "공문 주소", "공문주소")),
                "대리인": _clean(_get(row, m, "대리인")),
                "구조변경": _clean(_get(row, m, "구조변경")),
                "비고": _clean(_get(row, m, "비고")),
                "전화 메모": _clean(_get(row, m, "전화 메모", "전화메모")),
                "가입여부": membership,
                "부과구분": charge_item(membership),
            }
            rows.append(row_out)
            if preview_limit and len(rows) >= preview_limit:
                return rows
    return rows


def _latest_active_month(headers: list[str], rows_cache: list[tuple]) -> int:
    """실제 입력된 마지막 월의 미수금 컬럼을 찾는다.

    미수금 파일의 월별 미수금은 매월 말 누적 잔액이므로 모두 합산하면 안 된다.
    """
    month_indices: dict[int, int] = {}
    for i, h in enumerate(headers):
        m = re.search(r"(\d{1,2})월\s*미수금", _clean(h).replace(" ", ""))
        if m:
            month_indices[int(m.group(1))] = i
    latest = 0
    for month, idx in month_indices.items():
        for row in rows_cache[:800]:
            if idx < len(row) and _clean(row[idx]) not in {"", "-", "–", "—"}:
                latest = max(latest, month)
                break
    return latest or max(month_indices.keys() or [0])


def _iter_arrears_rows(file_bytes: bytes, preview_limit: int | None = None) -> list[dict[str, Any]]:
    wb = _read_workbook(file_bytes)
    if "2026년회비내역" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="미수금 파일에서 '2026년회비내역' 시트를 찾지 못했습니다.")
    ws = wb["2026년회비내역"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, max_col=80, values_only=True))
    headers = [_clean(x) for x in header_row]

    def col_idx(label: str) -> int | None:
        nl = _norm_col(label)
        for i, h in enumerate(headers):
            if _norm_col(h) == nl:
                return i
        return None

    base_cols = {
        "지역": col_idx("지역"), "계정": col_idx("계정"), "비고": col_idx("비고"),
        "차량번호": col_idx("차량번호"), "성명": col_idx("성명"), "대수": col_idx("대수"), "이월금": col_idx("이월금"),
    }
    month_cols = {month: col_idx(f"{month}월 미수금") for month in range(1, 13)}
    raw_rows = list(ws.iter_rows(min_row=2, max_col=80, values_only=True))
    latest_month = _latest_active_month(headers, raw_rows)
    latest_idx = month_cols.get(latest_month)

    rows: list[dict[str, Any]] = []
    for row in raw_rows:
        vehicle = _clean(row[base_cols["차량번호"]]) if base_cols["차량번호"] is not None else ""
        name = _person_name(row[base_cols["성명"]]) if base_cols["성명"] is not None else ""
        if not _valid_vehicle(vehicle) or not name:
            continue
        monthly_values: dict[str, int] = {}
        for month in range(1, 13):
            idx = month_cols.get(month)
            monthly_values[f"{month}월 미수금"] = _money(row[idx]) if idx is not None and idx < len(row) else 0
        current_amount = _money(row[latest_idx]) if latest_idx is not None and latest_idx < len(row) else 0
        out = {
            "지역": _clean(row[base_cols["지역"]]) if base_cols["지역"] is not None else "",
            "계정": _clean(row[base_cols["계정"]]) if base_cols["계정"] is not None else "",
            "비고": _clean(row[base_cols["비고"]]) if base_cols["비고"] is not None else "",
            "차량번호": vehicle,
            "성명": name,
            "대수": _clean(row[base_cols["대수"]]) if base_cols["대수"] is not None else "",
            "이월금": _money(row[base_cols["이월금"]]) if base_cols["이월금"] is not None else 0,
            "기준월": f"2026-{latest_month:02d}" if latest_month else "2026-00",
            "미수금액": current_amount,
            **monthly_values,
        }
        rows.append(out)
        if preview_limit and len(rows) >= preview_limit:
            return rows
    return rows

def _iter_deposit_rows(file_bytes: bytes, preview_limit: int | None = None) -> list[dict[str, Any]]:
    # 일반 통장거래내역은 형식이 다양하므로 pandas로 첫 시트 헤더 자동 추정
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, dtype=object)
    df = df.dropna(how="all").head(preview_limit or 5000)
    return [{str(k): _json(v) for k, v in r.items()} for r in df.to_dict("records")]


@router.post("/preview")
async def preview_import(file_type: str = Form(...), file: UploadFile = File(...)):
    data = await file.read()
    try:
        if file_type == "members":
            rows = _iter_license_rows(data, preview_limit=300)
            return {
                "ok": True,
                "filename": file.filename,
                "type": "members",
                "message": "전체면허자현황은 '개인' + '택배' 시트만 읽습니다. 업체세분/업체/차량집계는 제외됩니다.",
                "total_rows": len(rows),
                "columns": STANDARD_MEMBER_COLUMNS,
                "raw_columns": ["개인 시트", "택배 시트"],
                "sample": rows,
            }
        if file_type == "arrears":
            rows = _iter_arrears_rows(data, preview_limit=300)
            return {
                "ok": True,
                "filename": file.filename,
                "type": "arrears",
                "message": "미수금명단은 '2026년회비내역' 시트의 이월금 + 월별 미수금만 읽습니다.",
                "total_rows": len(rows),
                "columns": ARREARS_PREVIEW_COLUMNS,
                "raw_columns": ["2026년회비내역"],
                "sample": rows,
            }
        if file_type == "deposits":
            rows = _iter_deposit_rows(data, preview_limit=300)
            cols = list(rows[0].keys()) if rows else []
            return {"ok": True, "filename": file.filename, "type": "deposits", "total_rows": len(rows), "columns": cols, "raw_columns": cols, "sample": rows}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"엑셀 미리보기 실패: {str(exc)[:500]}") from exc
    raise HTTPException(status_code=400, detail="file_type은 members / arrears / deposits 중 하나여야 합니다.")


@router.post("/reset")
def reset_misu_data(db: Session = Depends(get_db)):
    try:
        # FK 순서 때문에 자식 테이블부터 비운다. misu_* 업무 테이블만 대상으로 한다.
        db.execute(delete(MemberHistory))
        db.execute(delete(Payment))
        db.execute(delete(ReceivableItem))
        db.execute(delete(Closure))
        db.execute(delete(Deposit))
        db.execute(delete(Pending))
        db.execute(delete(Member))
        db.commit()
        return {"ok": True, "message": "misu_* 업무 데이터 초기화 완료"}
    except SQLAlchemyError as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"초기화 실패: {str(exc)[:300]}") from exc


@router.post("/commit")
async def commit_import(file_type: str = Form(...), file: UploadFile = File(...), db: Session = Depends(get_db)):
    data = await file.read()
    if file_type == "members":
        rows = _iter_license_rows(data, preview_limit=None)
        inserted = updated = skipped = 0
        next_no = _current_max_member_no(db) + 1
        used_mgmt: set[str] = set(db.scalars(select(Member.mgmt_no)).all())
        for row in rows:
            vehicle = row["차량번호"]
            name = row["성명"]
            if not _valid_vehicle(vehicle) or not name:
                skipped += 1
                continue
            existing = _find_existing_member(db, name, vehicle)
            membership = row["가입여부"]
            item = charge_item(membership)
            cert_date = _parse_date(row.get("자격증명 발급일자"))
            join_date = _parse_date(row.get("가입일자"))
            start_ym = next_month_ym(join_date if membership == "협회가입" else cert_date)
            byear = _birth_year(row.get("주민등록번호", ""))
            age = (date.today().year - byear) if byear else None
            amount = monthly_charge(membership, age=age, birth_year=byear)
            memo_parts = []
            for k in ["주소", "사업자등록번호", "소속업체", "공문 주소", "대리인", "구조변경", "비고", "전화 메모"]:
                if row.get(k):
                    memo_parts.append(f"{k}:{row[k]}")
            memo = " / ".join(memo_parts)[:1000] or "엑셀 업로드 반영"
            if existing:
                m = existing
                m.mgmt_no = m.mgmt_no or _make_mgmt_no(row.get("관리번호", ""), next_no)
                m.reg_type = "양도양수" if str(m.mgmt_no).startswith("양") else "신규"
                m.name = name
                m.vehicle_no = vehicle
                m.phone = _clip(row.get("핸드폰") or row.get("전화번호"), 20)
                m.sigun = row.get("지역") or "미분류"
                m.region_raw = row.get("지역") or "미분류"
                m.member_type = row.get("회원구분") or "개인"
                m.membership = membership
                m.birth_year = byear
                m.cert_issue_date = cert_date
                m.assoc_join_date = join_date
                m.billing_start_ym = start_ym
                m.charge_item = item
                m.monthly_charge = amount
                m.status = m.status or "정상"
                m.cert_missing = cert_date is None
                m.memo = memo
                updated += 1
            else:
                mgmt = _make_mgmt_no(row.get("관리번호", ""), next_no)
                base = mgmt
                suffix = 2
                while mgmt in used_mgmt:
                    tail = f"-{suffix}"
                    mgmt = f"{base[:16-len(tail)]}{tail}"
                    suffix += 1
                used_mgmt.add(mgmt)
                m = Member(
                    id=_next_member_id_from_no(next_no),
                    mgmt_no=mgmt,
                    reg_type="양도양수" if mgmt.startswith("양") else "신규",
                    name=name,
                    vehicle_no=vehicle,
                    phone=_clip(row.get("핸드폰") or row.get("전화번호"), 20),
                    sigun=row.get("지역") or "미분류",
                    region_raw=row.get("지역") or "미분류",
                    member_type=row.get("회원구분") or "개인",
                    membership=membership,
                    birth_year=byear,
                    cert_issue_date=cert_date,
                    assoc_join_date=join_date,
                    billing_start_ym=start_ym,
                    charge_item=item,
                    monthly_charge=amount,
                    status="정상",
                    is_disconnected=False,
                    cert_missing=cert_date is None,
                    memo=memo,
                )
                db.add(m)
                next_no += 1
                inserted += 1
            if (inserted + updated) % 500 == 0:
                db.commit()
        db.commit()
        return {"ok": True, "filename": file.filename, "type": "members", "inserted": inserted, "updated": updated, "skipped": skipped, "errors": []}

    if file_type == "arrears":
        rows = _iter_arrears_rows(data, preview_limit=None)
        inserted = updated = skipped = 0
        unmatched: list[dict[str, str]] = []
        for row in rows:
            member = _find_member_for_arrears(db, row["성명"], row["차량번호"])
            if not member:
                skipped += 1
                if len(unmatched) < 30:
                    unmatched.append({"성명": row["성명"], "차량번호": row["차량번호"], "사유": "전체자명단에서 이름+차량번호 뒤4자리 매칭 실패"})
                continue

            # 중요: 월별 미수금은 누적 잔액이다.
            # 이월금 + 1월 미수금 + 2월 미수금 ... 을 합산하면 금액이 틀어진다.
            # 현재 미수금은 실제 입력된 마지막 월의 미수금액 1개만 저장한다.
            amt = _money(row.get("미수금액"))
            ym = row.get("기준월") or "2026-00"
            if amt <= 0:
                skipped += 1
                continue

            item = db.scalar(select(ReceivableItem).where(ReceivableItem.member_id == member.id, ReceivableItem.ym == ym).limit(1))
            if item:
                item.amount = amt
                item.charge_item = member.charge_item
                item.is_paid = False
                updated += 1
            else:
                db.add(ReceivableItem(member_id=member.id, ym=ym, charge_item=member.charge_item, amount=amt, is_paid=False))
                inserted += 1
            if (inserted + updated) % 500 == 0:
                db.commit()
        db.commit()
        return {"ok": True, "filename": file.filename, "type": "arrears", "inserted": inserted, "updated": updated, "skipped": skipped, "errors": unmatched}

    if file_type == "deposits":
        rows = _iter_deposit_rows(data, preview_limit=None)
        inserted = skipped = 0
        for r in rows:
            # 최소한 이름/금액처럼 보이는 컬럼을 찾아 저장
            keys = list(r.keys())
            name_key = next((k for k in keys if "입금자" in k or "예금주" in k or "거래기록" in k or "내용" in k), None)
            amt_key = next((k for k in keys if "입금액" in k or "금액" in k), None)
            date_key = next((k for k in keys if "일자" in k or "일" == k or "날짜" in k), None)
            name = _clean(r.get(name_key)) if name_key else ""
            amt = _money(r.get(amt_key)) if amt_key else 0
            if not name or amt <= 0:
                skipped += 1
                continue
            db.add(Deposit(deposit_date=_parse_date(r.get(date_key)) or date.today(), depositor_name=_clip(name, 40) or "입금자미상", amount=amt, memo="엑셀 업로드", status="대기", is_excluded=False))
            inserted += 1
        db.commit()
        return {"ok": True, "filename": file.filename, "type": "deposits", "inserted": inserted, "updated": 0, "skipped": skipped, "errors": []}

    raise HTTPException(status_code=400, detail="file_type은 members / arrears / deposits 중 하나여야 합니다.")
